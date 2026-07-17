#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
데이터 재생성 도구 — 앱(SCHOOLS)과 마스터 엑셀을 한 번에 맞춘다.

입력(있으면 병합, 없으면 무시):
  data/대표학과_미배정_76개교_배정용.csv   ← '대표학과(여기에_작성)' 열을 채워서 저장
  data/학교_주소좌표_입력용.csv            ← 도로명주소/위도/경도 열을 채워서 저장

동작:
  - app/홍보배차판.jsx 의 SCHOOLS 배열을 읽어
  - 대표학과(d)·주소(addr)·좌표(lat/lng)를 병합하고
  - 우선순위점수(sc)·등급(t)을 공식대로 재계산한 뒤
  - --write 시 jsx의 SCHOOLS를 갱신하고, 마스터 엑셀의 대표학과 칸을 채운다.

개인정보(담당교사 등)는 절대 다루지 않는다. 학교 단위 데이터만.

사용:
  python3 scripts/rebuild_data.py            # 미리보기(변경사항만 출력)
  python3 scripts/rebuild_data.py --write     # 실제 반영
"""
import json, re, sys, csv, os
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
JSX = ROOT / "app" / "홍보배차판.jsx"
XLSX = ROOT / "data" / "홍보대상학교_마스터.xlsx"
DEPT_CSV = ROOT / "data" / "대표학과_미배정_76개교_배정용.csv"
GEO_CSV = ROOT / "data" / "학교_주소좌표_입력용.csv"

# 등급 경계값 (안내 시트 B7/B8과 동일하게 유지)
TIER_HIGH = 17.0
TIER_MID = 8.6


def num(v):
    return 0.0 if v is None else float(v)


def score(s):
    return round(num(s.get("y3")) * 2 + num(s.get("y26")) * 3 + num(s.get("c3")) * 0.6, 1)


def tier(sc):
    return "상" if sc >= TIER_HIGH else ("중" if sc >= TIER_MID else "하")


def load_schools():
    src = JSX.read_text(encoding="utf-8")
    m = re.search(r"const SCHOOLS = (\[.*?\]);", src, re.S)
    if not m:
        sys.exit("SCHOOLS 배열을 찾지 못했습니다.")
    return src, m, json.loads(m.group(1))


def read_csv(path):
    if not path.exists():
        return []
    with open(path, encoding="utf-8-sig", newline="") as f:
        return list(csv.DictReader(f))


def find_col(row, *cands):
    for c in cands:
        for k in row.keys():
            if k and c in k:
                return k
    return None


def key(n, g):
    return f"{n}|{g}"


def main():
    write = "--write" in sys.argv
    src, m, schools = load_schools()

    # 대표학과 병합
    dept = {}
    for r in read_csv(DEPT_CSV):
        kcol = find_col(r, "학교명")
        gcol = find_col(r, "지역")
        dcol = find_col(r, "대표학과")
        if kcol and dcol and r.get(dcol, "").strip():
            dept[key(r[kcol].strip(), (r.get(gcol) or "").strip())] = r[dcol].strip()

    # 주소/좌표 병합
    geo = {}
    for r in read_csv(GEO_CSV):
        kcol = find_col(r, "학교명")
        gcol = find_col(r, "지역")
        acol = find_col(r, "주소")
        latc = find_col(r, "위도", "lat")
        lngc = find_col(r, "경도", "lng")
        if not kcol:
            continue
        k = key(r[kcol].strip(), (r.get(gcol) or "").strip())
        entry = {}
        if acol and (r.get(acol) or "").strip():
            entry["addr"] = r[acol].strip()
        try:
            if latc and (r.get(latc) or "").strip():
                entry["lat"] = round(float(r[latc]), 6)
            if lngc and (r.get(lngc) or "").strip():
                entry["lng"] = round(float(r[lngc]), 6)
        except ValueError:
            pass
        if entry:
            geo[k] = entry

    changes = []
    for s in schools:
        k = key(s["n"], s["g"])
        if k in dept and not s.get("d"):
            changes.append(f"  대표학과  {s['n']}({s['g']}) : (없음) → {dept[k]}")
            s["d"] = dept[k]
        if k in geo:
            for f in ("addr", "lat", "lng"):
                if f in geo[k] and s.get(f) != geo[k][f]:
                    changes.append(f"  {f:5} {s['n']}({s['g']}) : {s.get(f)} → {geo[k][f]}")
                    s[f] = geo[k][f]
        new_sc = score(s)
        if s.get("sc") != new_sc:
            changes.append(f"  점수    {s['n']}({s['g']}) : {s.get('sc')} → {new_sc}")
            s["sc"] = new_sc
        s["t"] = tier(s["sc"])

    blanks = sum(1 for s in schools if not s.get("d"))
    print(f"학교 {len(schools)}개 · 대표학과 미배정 {blanks}개 · 변경 {len(changes)}건")
    for c in changes[:80]:
        print(c)
    if len(changes) > 80:
        print(f"  … 외 {len(changes) - 80}건")

    if not write:
        print("\n(미리보기입니다. 실제 반영하려면 --write 를 붙이세요.)")
        return

    # jsx 갱신: SCHOOLS를 한 줄 JSON으로 직렬화(기존 스타일 유지)
    keys_order = ["n", "g", "s", "y3", "y26", "c3", "st", "d", "sc", "t", "addr", "lat", "lng"]
    def ser(s):
        parts = []
        for kk in keys_order:
            if kk in s and s[kk] is not None:
                parts.append(f"{json.dumps(kk, ensure_ascii=False)}: {json.dumps(s[kk], ensure_ascii=False)}")
        return "{" + ", ".join(parts) + "}"
    arr = "[" + ", ".join(ser(s) for s in schools) + "]"
    JSX.write_text(src[:m.start(1)] + arr + src[m.end(1):], encoding="utf-8")
    print(f"\n✔ {JSX} 의 SCHOOLS 갱신 완료")

    # 마스터 엑셀 대표학과 칸 채우기(있을 때, 서식/수식 보존)
    if XLSX.exists():
        try:
            import openpyxl
            wb = openpyxl.load_workbook(XLSX)
            ws = wb["홍보대상학교"] if "홍보대상학교" in wb.sheetnames else wb.active
            head = {ws.cell(1, c).value: c for c in range(1, ws.max_column + 1)}
            cN = next((c for h, c in head.items() if h and "학교명" in str(h)), None)
            cG = next((c for h, c in head.items() if h and "지역" in str(h)), None)
            cD = next((c for h, c in head.items() if h and "대표학과" in str(h)), None)
            dmap = {key(s["n"], s["g"]): s.get("d") for s in schools}
            filled = 0
            if cN and cD:
                for r in range(2, ws.max_row + 1):
                    n = ws.cell(r, cN).value
                    if not n:
                        continue
                    g = ws.cell(r, cG).value if cG else ""
                    d = dmap.get(key(str(n).strip(), str(g or "").strip()))
                    if d and not ws.cell(r, cD).value:
                        ws.cell(r, cD).value = d
                        filled += 1
            wb.save(XLSX)
            print(f"✔ {XLSX} 대표학과 {filled}칸 반영 완료")
        except ImportError:
            print("! openpyxl 미설치 → 엑셀 갱신 건너뜀 (pip install openpyxl)")


if __name__ == "__main__":
    main()
